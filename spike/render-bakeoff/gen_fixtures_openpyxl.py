"""Fixtures for features the asp 0.14 write surface cannot express.

asp has no canonical write op for row heights or for hiding rows/columns:
`asp write batch column-size` only carries {auto|width}, `sheet-layout` only carries
freeze/zoom/gridlines/print, and `structure` only insert/delete/merge/move.
These two fixtures are therefore produced with openpyxl.
"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

out = sys.argv[1]
os.makedirs(out, exist_ok=True)

# gen_08_rowheights
wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
ws["A1"] = "Row heights (points)"
ws["A1"].font = Font(bold=True, size=14)
for i, h in enumerate([8, 12, 15, 20, 30, 45, 60, 90]):
    r = 3 + i
    ws.cell(r, 1, f"{h}pt").font = Font(size=min(h * 0.6, 22))
    ws.cell(r, 2, "Ag baseline check")
    ws.cell(r, 3, r * 111)
    ws.row_dimensions[r].height = h
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 24
ws["A12"] = "wrapped, auto height"
ws["B12"] = "A long wrapped sentence in a cell whose height Excel computed automatically."
ws["B12"].alignment = Alignment(wrap_text=True, vertical="top")
wb.save(os.path.join(out, "gen_08_rowheights.xlsx"))

# gen_09_hidden
wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
ws["A1"] = "Hidden rows and columns"
ws["A1"].font = Font(bold=True, size=14)
for r in range(3, 16):
    for c in range(1, 9):
        ws.cell(r, c, f"{chr(64+c)}{r}")
for r in (5, 6, 10):
    ws.row_dimensions[r].hidden = True
for c in ("C", "D", "G"):
    ws.column_dimensions[c].hidden = True
ws.column_dimensions["B"].width = 20
ws["A17"] = "rows 5,6,10 hidden; cols C,D,G hidden"
fill = PatternFill("solid", fgColor="FFFFE699")
thin = Side(style="thin", color="FF808080")
for c in range(1, 9):
    ws.cell(3, c).fill = fill
    ws.cell(3, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
wb.save(os.path.join(out, "gen_09_hidden.xlsx"))
print("ok gen_08_rowheights gen_09_hidden")
