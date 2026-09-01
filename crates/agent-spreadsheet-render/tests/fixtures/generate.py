#!/usr/bin/env python3
"""Generate the fixtures the copied bake-off corpus does not cover.

The canonical `asp` write surface has no operation for conditional formatting,
charts, images, text rotation or hidden rows/columns, so these are authored
with openpyxl. That gap is recorded in `docs/architecture/renderer-bake-off.md`
and a follow-up ticket adds the missing write ops.

    UV_CACHE_DIR=$TMPDIR/uvcache uv run --with openpyxl --with pillow \
        python3 crates/agent-spreadsheet-render/tests/fixtures/generate.py

`gen_11_dashboard.xlsx` additionally goes through the product's own
recalculation so its formulas carry cached values (the renderer never
recalculates):

    cargo build -p agent-spreadsheet --bin asp --config 'build.rustc-wrapper=""'
    ./target/debug/asp workbook recalculate gen_11_dashboard.xlsx \
        --output gen_11_dashboard.recalc.xlsx --force
"""

import pathlib
import posixpath
import re
import shutil
import struct
import tempfile
import zipfile
import zlib

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill

HERE = pathlib.Path(__file__).parent


def excelize(path):
    """Make an openpyxl package look like one Excel wrote.

    Rewrite absolute relationship targets to package-relative ones.

    openpyxl writes `Target="/xl/drawings/drawing1.xml"`. Excel writes
    `Target="../drawings/drawing1.xml"`, and umya-spreadsheet only resolves the
    relative form — it drops the drawing entirely for the absolute one, so both
    `chart_omitted` and `image_omitted` would be untestable. Rewriting here
    makes the fixture match what Excel itself produces.
    """
    with zipfile.ZipFile(path) as archive:
        entries = [(item, archive.read(item.filename)) for item in archive.infolist()]
    changed = []
    for item, data in entries:
        if not item.filename.endswith(".rels"):
            continue
        text = data.decode("utf-8")
        base = posixpath.dirname(posixpath.dirname(item.filename))
        def fix(match_text):
            target = match_text
            if not target.startswith("/"):
                return target
            return posixpath.relpath(target.lstrip("/"), base or ".")
        out = []
        for chunk in text.split('Target="'):
            if not out:
                out.append(chunk)
                continue
            value, rest = chunk.split('"', 1)
            out.append(fix(value) + '"' + rest)
        text = 'Target="'.join(out)
        changed.append((item.filename, text.encode("utf-8")))
    for item, data in entries:
        if item.filename.startswith("xl/drawings/drawing"):
            rewritten = prefix_default_namespace(
                data.decode("utf-8"), "xdr", SPREADSHEET_DRAWING_NS)
            changed.append((item.filename, rewritten.encode("utf-8")))
        elif item.filename.startswith("xl/charts/chart"):
            rewritten = prefix_default_namespace(data.decode("utf-8"), "c", CHART_NS)
            changed.append((item.filename, rewritten.encode("utf-8")))
    if not changed:
        return
    replacements = dict(changed)
    handle, temporary = tempfile.mkstemp(suffix=".xlsx")
    import os
    os.close(handle)
    with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as out:
        for item, data in entries:
            out.writestr(item, replacements.get(item.filename, data))
    shutil.move(temporary, path)


SPREADSHEET_DRAWING_NS = (
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
)
CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"


def prefix_default_namespace(xml, prefix, namespace):
    """Rewrite a default-namespaced part into the prefixed form Excel writes.

    Excel writes `<xdr:wsDr xmlns:xdr="...">` and `<c:chartSpace
    xmlns:c="...">`; openpyxl writes both with a default `xmlns`.
    umya-spreadsheet matches element names literally (`b"xdr:wsDr"` in
    `reader/xlsx/drawing.rs`, `b"c:chart"` and friends in the chart reader), so
    it reads the Excel form and silently ignores the openpyxl one — which would
    leave `chart_omitted` and `image_omitted` unreachable. Rewriting to Excel's
    form is what makes these fixtures represent files the product is actually
    handed.
    """
    xml = xml.replace(f' xmlns="{namespace}"', "")
    xml = re.sub(r"<(/?)([A-Za-z][A-Za-z0-9]*)([ />])", rf"<\1{prefix}:\2\3", xml)
    root = re.search(rf"<{prefix}:([A-Za-z0-9]+)", xml).group(1)
    return xml.replace(
        f"<{prefix}:{root}",
        f'<{prefix}:{root} xmlns:{prefix}="{namespace}"',
        1,
    )


def conditional_format():
    """conditional_format_omitted: rules are read but never evaluated."""
    book = Workbook()
    sheet = book.active
    sheet.title = "CondFmt"
    sheet["A1"] = "Region"
    sheet["B1"] = "Variance"
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    for index, (region, variance) in enumerate(
        [("North", 0.12), ("South", -0.08), ("East", 0.31), ("West", -0.22)],
        start=2,
    ):
        sheet.cell(row=index, column=1, value=region)
        cell = sheet.cell(row=index, column=2, value=variance)
        cell.number_format = "0.0%;[Red]-0.0%"
    sheet.conditional_formatting.add(
        "B2:B5",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill(
            start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")),
    )
    sheet.conditional_formatting.add(
        "B2:B5", ColorScaleRule(start_type="min", start_color="FFF8696B",
                                end_type="max", end_color="FF63BE7B"))
    book.save(HERE / "cf_conditional_format.xlsx")


def chart():
    """chart_omitted: the data table renders, the chart does not."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Chart"
    sheet.append(["Quarter", "Revenue"])
    for row in [("Q1", 1200), ("Q2", 1580), ("Q3", 1410), ("Q4", 1930)]:
        sheet.append(row)
    chart = BarChart()
    chart.title = "Revenue by quarter"
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=5))
    # A two-cell anchor, which is what Excel writes for a chart. umya only
    # classifies a graphic frame as a chart inside `xdr:twoCellAnchor`; a chart
    # in openpyxl's default `oneCellAnchor` is dropped on the floor.
    chart.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=3, row=1), to=AnchorMarker(col=9, row=15)
    )
    sheet.add_chart(chart)
    book.save(HERE / "chart_bar.xlsx")
    excelize(HERE / "chart_bar.xlsx")


def unicode_wide():
    """font_substituted plus the .notdef path.

    Row 2 is inside the Carlito subset and must render as glyphs; rows 3 to 6
    are outside it and must render as visible .notdef boxes, never as blanks.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Unicode"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 32
    rows = [
        ("script", "sample"),
        ("latin-1 + currency", "æøå ß £ € ¥ ₹ ½ ±"),
        ("greek", "α β γ Ω"),
        ("cyrillic", "а б в Я"),
        ("cjk", "日本語 中文"),
        ("hebrew", "א ב ג"),
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    book.save(HERE / "unicode_wide.xlsx")


def uncached_formulas():
    """formulas_unevaluated: openpyxl writes formulas with no cached value."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Formulas"
    sheet["A1"] = "input"
    sheet["A2"] = 10
    sheet["A3"] = 32
    sheet["B1"] = "total"
    sheet["B2"] = "=SUM(A2:A3)"
    sheet["B3"] = "=B2*2"
    book.save(HERE / "formulas_uncached.xlsx")


def _tiny_png(path):
    """A 4x4 red PNG, written by hand so the fixture needs no image library."""
    def chunk(kind, payload):
        body = kind + payload
        return struct.pack(">I", len(payload)) + body + struct.pack(
            ">I", zlib.crc32(body) & 0xFFFFFFFF)

    raw = b"".join(b"\x00" + b"\xff\x00\x00\xff" * 4 for _ in range(4))
    data = (
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", struct.pack(">IIBBBBB", 4, 4, 8, 6, 0, 0, 0))
        + chunk(b"IDAT", zlib.compress(raw))
        + chunk(b"IEND", b"")
    )
    path.write_bytes(data)
    return path


def warnings_grabbag():
    """image_omitted, text_rotation_omitted, pattern_fill_approximated and
    number_format_approximated in one small workbook."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Warnings"
    sheet["A1"] = "rotated"
    sheet["A1"].alignment = Alignment(textRotation=45)
    sheet["B1"] = "hatched"
    sheet["B1"].fill = PatternFill(
        patternType="lightUp", start_color="FF735773", end_color="FFFFFFFF")
    sheet["C1"] = 45000
    sheet["C1"].number_format = "yyyy-mm-dd"
    png = _tiny_png(HERE / "_tiny.png")
    try:
        sheet.add_image(Image(str(png)), "E2")
        book.save(HERE / "warnings_grabbag.xlsx")
        excelize(HERE / "warnings_grabbag.xlsx")
    finally:
        png.unlink()


if __name__ == "__main__":
    conditional_format()
    chart()
    unicode_wide()
    uncached_formulas()
    warnings_grabbag()
    print("wrote fixtures to", HERE)
